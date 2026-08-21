from __future__ import annotations

import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
coverage = json.loads((ROOT / "storage/feature_coverage.json").read_text())

wb = Workbook()
ws = wb.active
ws.title = "Overview"

header_fill = PatternFill("solid", fgColor="111111")
accent_fill = PatternFill("solid", fgColor="DFF7FB")
header_font = Font(color="FFFFFF", bold=True)
subheader_font = Font(bold=True, color="007E9E")


def write_table(sheet, start_row, headers, rows):
    for col, value in enumerate(headers, 1):
        cell = sheet.cell(start_row, col, value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_offset, row in enumerate(rows, 1):
        for col, value in enumerate(row, 1):
            cell = sheet.cell(start_row + r_offset, col, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    return start_row + len(rows)


def style_sheet(sheet):
    sheet.freeze_panes = "A2"
    for col_cells in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, min(52, len(str(cell.value))))
        sheet.column_dimensions[col_letter].width = max(12, max_len + 2)

ws["A1"] = "Universe Civilization: Empire at Wars Feature and Contract Workbook"
ws["A1"].font = Font(size=16, bold=True, color="007E9E")
ws["A3"] = "Generated UTC"
ws["B3"] = coverage["generated_at"]
ws["A4"] = "Registered routes"
ws["B4"] = coverage["routes"]
ws["A5"] = "Complete route support sets"
ws["B5"] = coverage["complete_routes"]
ws["A6"] = "Incomplete route support sets"
ws["B6"] = coverage["incomplete_routes"]
ws["A7"] = "PHP page files"
ws["B7"] = coverage["support_files"]["php_pages"]
ws["A8"] = "Service classes"
ws["B8"] = coverage["support_files"]["services"]
ws["A9"] = "SQL migration/schema files"
ws["B9"] = coverage["support_files"]["sql_migrations"]
ws["A10"] = "Test files"
ws["B10"] = coverage["support_files"]["tests"]
for row in range(3, 11):
    ws.cell(row, 1).font = subheader_font

routes = []
contracts = []
for item in coverage["rows"]:
    routes.append([
        item["group"], item["route"], item["title"], "COMPLETE" if item["complete"] else "INCOMPLETE",
        ", ".join(item["actions"]), ", ".join(item["tables"]), ", ".join(item["feedback_states"]),
        ", ".join(item["missing"]),
    ])
    contracts.append([
        item["route"], item["title"], ", ".join(item["actions"]) or "read-only", ", ".join(item["tables"]),
        ", ".join(item["feedback_states"]) or "route-defined / renderer-defined",
        "authenticated; CSRF; RBAC; ownership; server-side validation; transaction where mutating",
    ])

ws_routes = wb.create_sheet("Route Coverage")
write_table(ws_routes, 1, ["Group", "Route", "Title", "Support", "Actions", "Data Sources", "Feedback States", "Missing Files"], routes)
style_sheet(ws_routes)

ws_contracts = wb.create_sheet("Backend Contracts")
write_table(ws_contracts, 1, ["Route", "Title", "Actions / Intents", "Tables / Data Sources", "Feedback States", "Security Contract"], contracts)
style_sheet(ws_contracts)

sql_objects = []
for sql_file in sorted((ROOT / "sql").glob("*.sql")):
    text = sql_file.read_text(errors="ignore")
    for match in re.finditer(r"CREATE\\s+(?:TABLE|VIEW)\\s+(?:IF\\s+NOT\\s+EXISTS\\s+)?`?([A-Za-z0-9_]+)`?", text, re.I):
        sql_objects.append([sql_file.name, match.group(1), "table/view", "forward migration/schema"])
ws_db = wb.create_sheet("Database Objects")
write_table(ws_db, 1, ["Source File", "Object", "Type", "Purpose"], sql_objects)
style_sheet(ws_db)

mechanics = [
    ["Economy", "8 core resources", "Metal, Crystal, Naquadah, Energy, Dark Matter, Food, Water, Population", "player_resources; EconomyService"],
    ["Population", "Capacity and life support", "Population capacity, food/water consumption, growth, morale", "player_resources; player_colonies; EconomyService"],
    ["Progression", "21 tiers × 23 levels", "Tier mastery, level effects, ascension eligibility", "progression_states; ProgressionService; WorldService"],
    ["Combat", "Deterministic force resolution", "Validated force comparison, technology/race modifiers, protection, casualties, loot", "GameService::resolveCombat; battles; battle_reports"],
    ["Covert", "Recon, spy, sabotage", "Detection, bounded disruption, probability caps, agent expenditure", "OGameService/WorldService; covert tables; intelligence_reports"],
    ["Fleet", "Fleet missions and turn settlement", "Queued missions, cooldowns, resource and ownership checks", "fleet_missions; player_cooldowns; GameService/OGameService"],
    ["Colonies", "Exploration and colonization", "Habitability, occupancy, colony capacity, biome and planet bonuses", "PlanetService; WorldService; player_colonies"],
    ["Mothership", "Modules and exploration", "Hull/modules, science, distance, anomaly and rarity yields", "MothershipService; MothershipExplorationService"],
    ["Markets", "Resource, weapon, mercenary exchange", "Validated listings, fees, expiry, balances and ownership", "ResourceMarketService; WeaponMarketService; MercenaryMarketService"],
    ["Security", "Transactional server authority", "Authentication, CSRF, RBAC, ownership, cooldowns, rollback-safe writes", "auth.php; actions/game.php; service layer"],
]
ws_mech = wb.create_sheet("Mechanics Matrix")
write_table(ws_mech, 1, ["Module", "Mechanic", "Implemented Behavior", "Primary Contracts"], mechanics)
style_sheet(ws_mech)

tests = []
for test_file in sorted((ROOT / "tests").glob("*.php")):
    name = test_file.stem
    category = "integration" if "contract" in name or name in {"smoke_test", "page_modules_integration", "turn_settlement_e2e"} else "unit/load/other"
    tests.append([test_file.name, category, "PHP test present", "Run with php tests/" + test_file.name])
ws_tests = wb.create_sheet("Test Inventory")
write_table(ws_tests, 1, ["Test File", "Category", "Coverage", "Command"], tests)
style_sheet(ws_tests)

for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.row == 1 and sheet.title != "Overview":
                cell.fill = header_fill
                cell.font = header_font
    sheet.sheet_view.showGridLines = False

out = ROOT / "docs" / "UNIVERSE CIVILIZATION: EMPIRE AT WARS_FEATURE_WORKBOOK.xlsx"
wb.save(out)
print(out)
